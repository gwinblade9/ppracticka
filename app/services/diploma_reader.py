from datetime import datetime
from pathlib import Path
import re

from openpyxl import load_workbook


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return format_date(value)
    return str(value).replace("\n", " ").strip()


def format_date(value: datetime) -> str:
    months = {
        1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня",
        7: "июля", 8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря",
    }
    return f"{value.day:02d} {months[value.month]} {value.year} года"


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", clean(value)).strip()


def parse_period_and_hours(text: str) -> dict:
    text = clean(text)
    result = {"Период": "", "Часы": "", "Дата_начала": "", "Дата_окончания": ""}

    period_match = re.search(r"с\s+(.+?)\s+по\s+(.+?)(?:,|$)", text, flags=re.IGNORECASE)
    if period_match:
        start = period_match.group(1).strip()
        end = period_match.group(2).strip()
        result["Дата_начала"] = start
        result["Дата_окончания"] = end
        result["Период"] = f"с {start} по {end}"

    hours_match = re.search(r"об[ъь]ем\s+([0-9]+)", text, flags=re.IGNORECASE)
    if hours_match:
        result["Часы"] = hours_match.group(1)

    return result


def find_header_row(rows):
    for index, row in enumerate(rows):
        row_text = [normalize_header(value).lower() for value in row]
        if "фамилия" in row_text and "имя" in row_text and "отчество" in row_text:
            return index
    raise ValueError("Не найдена строка заголовков. В таблице должны быть колонки Фамилия, Имя, Отчество")


def read_diploma_excel(file_path: Path) -> list[dict]:
    """Читает Excel с итоговой ведомостью для дипломов.

    Поддерживает таблицы, где сверху есть название программы и период обучения,
    а строка заголовков находится ниже.
    """
    workbook = load_workbook(file_path, data_only=True)
    students: list[dict] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        try:
            header_index = find_header_row(rows)
        except ValueError:
            continue

        # В примере название программы находится выше заголовков, чаще всего в колонке D.
        course_name = ""
        period_text = ""
        for row in rows[:header_index]:
            for value in row:
                text = clean(value)
                low = text.lower()
                if not course_name and text and "период" not in low and "ведомость" not in low and not text.isdigit():
                    # Берём первую длинную строку до строки заголовков как название программы.
                    if len(text) > 10:
                        course_name = text
                if "период обучения" in low or "объем" in low or "объём" in low:
                    period_text = text

        period_info = parse_period_and_hours(period_text)

        headers = [normalize_header(value) for value in rows[header_index]]
        for row in rows[header_index + 1:]:
            values = [clean(value) for value in row]
            if not any(values):
                continue

            item = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers)) if headers[i]}

            # Пропускаем технические строки без ФИО.
            if not clean(item.get("Фамилия")) or not clean(item.get("Имя")):
                continue

            item["id"] = len(students) + 1
            item["Лист"] = sheet.title
            item["Название_курса"] = item.get("Название_курса") or item.get("Специальность") or course_name
            item["Период"] = item.get("Период") or period_info["Период"]
            item["Дата_начала"] = item.get("Дата_начала") or period_info["Дата_начала"]
            item["Дата_окончания"] = item.get("Дата_окончания") or period_info["Дата_окончания"]
            item["Часы"] = item.get("Часы") or period_info["Часы"]

            # В бланке диплома дата решения комиссии совпадает с датой аттестационной комиссии.
            if not item.get("Дата_решения"):
                item["Дата_решения"] = item.get("Дата проведения аттестационной комиссии", "")

            students.append(item)

    if not students:
        raise ValueError("Не удалось прочитать студентов из Excel-файла для дипломов")

    return students
