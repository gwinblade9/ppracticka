from pathlib import Path
from openpyxl import load_workbook


def read_excel_table(file_path: Path) -> list[dict]:
    """Читает первый лист Excel-файла.

    Первая строка считается заголовками.
    Пустые строки пропускаются.
    """
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel-файл пустой")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]

    students = []
    for row in rows[1:]:
        values = [str(value).strip() if value is not None else "" for value in row]

        if not any(values):
            continue

        student = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        student["id"] = len(students) + 1
        students.append(student)

    return students
